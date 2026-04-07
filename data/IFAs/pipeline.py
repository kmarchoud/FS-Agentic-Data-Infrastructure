#!/usr/bin/env python3
"""
IFA Intelligence Pipeline — Keyridge Asset Management
Builds a comprehensive database of UK IFA/adviser/DFM firms.

Usage:
  python3 pipeline.py --test          # 100-firm test batch
  python3 pipeline.py --full          # Full run (all firms)
  python3 pipeline.py --resume        # Resume from last checkpoint
  python3 pipeline.py --phase N       # Run specific phase only
"""

import csv
import json
import os
import sys
import time
import argparse
import logging
import urllib.request
import urllib.error
import urllib.parse
import base64
from datetime import datetime
from pathlib import Path

# ── Configuration ─────────────────────────────────────────────────────────────

BASE_DIR = Path("/Users/karimmarchoud/Desktop/AI/FSInfra/Keyridge/Data/IFAs")
RAW_DIR = BASE_DIR / "raw"
ENRICHED_DIR = BASE_DIR / "enriched"
CHECKPOINT_DIR = BASE_DIR / "checkpoints"
LOG_DIR = BASE_DIR / "logs"

FCA_CSV_PATH = Path("/Users/karimmarchoud/Desktop/AI/FSInfra/Keyridge/Data/FCA/Investment Firms Register.csv")
PIMFA_MEMBERS_PATH = Path("/Users/karimmarchoud/Desktop/AI/FSInfra/Keyridge/Data/PIMFA/Member List.xlsx")
PIMFA_ASSOC_PATH = Path("/Users/karimmarchoud/Desktop/AI/FSInfra/Keyridge/Data/PIMFA/Associated Members List.xlsx")

# FCA API credentials
FCA_API_EMAIL = "karim.marchoud@regulex.io"
FCA_API_KEY = "e6acc1c947a4586da5ca619d45c03f96"
FCA_API_BASE = "https://register.fca.org.uk/services/V0.1"

# Companies House API key (set via --ch-key argument or CH_API_KEY env var)
CH_API_KEY = os.environ.get("CH_API_KEY", "")

# Rate limits (seconds between requests)
RATE_FCA = 1.0
RATE_CH = 1.0
RATE_WEB = 3.0

CHECKPOINT_FILE = CHECKPOINT_DIR / "master.json"
TEST_BATCH_SIZE = 100

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "pipeline.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("ifa_pipeline")

# ── Checkpoint System ─────────────────────────────────────────────────────────

def load_checkpoint():
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE) as f:
            return json.load(f)
    return {
        "phase": 0,
        "processed_frns": [],
        "processed_sources": [],
        "firm_count": 0,
        "last_updated": None,
        "phase_status": {},
    }


def save_checkpoint(state):
    state["last_updated"] = datetime.now().isoformat()
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(state, f, indent=2)


# ── Utilities ─────────────────────────────────────────────────────────────────

def jsonl_append(filepath, record):
    """Append a single JSON record to a .jsonl file."""
    with open(filepath, "a") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def jsonl_read(filepath):
    """Read all records from a .jsonl file."""
    records = []
    if not filepath.exists():
        return records
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def log_error(frn, source, phase, error_type, error_message):
    """Log an error to the errors file."""
    jsonl_append(LOG_DIR / "errors.jsonl", {
        "frn": frn,
        "source": source,
        "phase": phase,
        "error_type": error_type,
        "error_message": str(error_message)[:500],
        "timestamp": datetime.now().isoformat(),
    })


def fca_api_call(endpoint, retries=2):
    """Make a rate-limited FCA API call using curl (FCA API is case-sensitive on headers)."""
    import subprocess
    url = f"{FCA_API_BASE}/{endpoint}"

    for attempt in range(retries + 1):
        try:
            time.sleep(RATE_FCA)
            result = subprocess.run(
                [
                    "curl", "-s", "-X", "GET", url,
                    "-H", f"x-auth-email: {FCA_API_EMAIL}",
                    "-H", f"x-auth-key: {FCA_API_KEY}",
                    "-H", "Content-Type: application/json",
                ],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode != 0:
                log_error(endpoint, "fca_api", "api_call", "curl_error", result.stderr[:300])
                if attempt == retries:
                    return None
                continue

            if not result.stdout.strip():
                return None

            data = json.loads(result.stdout)

            # Check for API-level errors
            status = data.get("Status", "")
            if "01-11" in status:  # Unauthorised
                log_error(endpoint, "fca_api", "api_call", "unauthorised", status)
                return None
            if "21" in status and "Not Found" in data.get("Message", ""):
                return None  # Not found is not an error

            return data

        except json.JSONDecodeError as e:
            log_error(endpoint, "fca_api", "api_call", "json_decode", str(e))
            if attempt == retries:
                return None
        except subprocess.TimeoutExpired:
            log_error(endpoint, "fca_api", "api_call", "timeout", "curl timed out after 30s")
            if attempt == retries:
                return None
        except Exception as e:
            log_error(endpoint, "fca_api", "api_call", type(e).__name__, str(e))
            if attempt == retries:
                return None
            time.sleep(5)

    return None


def ch_api_call(endpoint, retries=2):
    """Make a rate-limited Companies House API call."""
    if not CH_API_KEY:
        return None
    url = f"https://api.company-information.service.gov.uk{endpoint}"
    auth = base64.b64encode(f"{CH_API_KEY}:".encode()).decode()
    headers = {
        "Authorization": f"Basic {auth}",
        "Accept": "application/json",
    }
    req = urllib.request.Request(url, headers=headers)

    for attempt in range(retries + 1):
        try:
            time.sleep(RATE_CH)
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code == 429:
                log.warning(f"CH rate limited on {endpoint}, backing off 60s...")
                time.sleep(60)
                continue
            elif e.code == 404:
                return None
            else:
                log_error(endpoint, "companies_house", "api_call", f"HTTP_{e.code}", str(e))
                if attempt == retries:
                    return None
        except Exception as e:
            log_error(endpoint, "companies_house", "api_call", type(e).__name__, str(e))
            if attempt == retries:
                return None
            time.sleep(5)

    return None


# ── PHASE 1: Universe Discovery ──────────────────────────────────────────────

def phase1_fca_csv():
    """Load and filter the FCA Investment Firms Register CSV."""
    log.info("Phase 1.1: Loading FCA Investment Firms Register CSV...")
    output_file = RAW_DIR / "source_fca_csv.jsonl"

    # Clear existing output
    if output_file.exists():
        output_file.unlink()

    firms = []
    with open(FCA_CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            frn = row.get("FRN", "").strip()
            name = row.get("Organisation Name", "").strip()
            status = row.get("Status", "").strip()
            auth_date = row.get("First Authorisation Date", "").strip()

            if status != "Authorised":
                continue
            if not frn or not name:
                continue

            record = {
                "frn": frn,
                "firm_name": name,
                "status": status,
                "authorisation_date": auth_date[:10] if auth_date else None,
                "lei": row.get("Legal Entity Identifier (LEI)", "").strip() or None,
                "source": "fca_csv",
            }
            firms.append(record)
            jsonl_append(output_file, record)

    log.info(f"Phase 1.1 complete: {len(firms)} authorised firms from FCA CSV")
    return firms


def phase1_pimfa():
    """Load PIMFA member and associated member lists."""
    log.info("Phase 1.2: Loading PIMFA member lists...")
    output_file = RAW_DIR / "source_pimfa.jsonl"

    if output_file.exists():
        output_file.unlink()

    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed. Run: pip3 install openpyxl")
        return []

    firms = []

    # Members
    if PIMFA_MEMBERS_PATH.exists():
        wb = openpyxl.load_workbook(PIMFA_MEMBERS_PATH, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row and row[0] else None
            if name and str(name).strip():
                record = {
                    "firm_name": str(name).strip(),
                    "pimfa_member": True,
                    "pimfa_type": "member",
                    "source": "pimfa",
                }
                firms.append(record)
                jsonl_append(output_file, record)
        wb.close()

    # Associated Members
    if PIMFA_ASSOC_PATH.exists():
        wb = openpyxl.load_workbook(PIMFA_ASSOC_PATH, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row and row[0] else None
            if name and str(name).strip():
                record = {
                    "firm_name": str(name).strip(),
                    "pimfa_member": True,
                    "pimfa_type": "associated_member",
                    "source": "pimfa",
                }
                firms.append(record)
                jsonl_append(output_file, record)
        wb.close()

    log.info(f"Phase 1.2 complete: {len(firms)} PIMFA members/associates loaded")
    return firms


def phase1_fca_api_ars(fca_firms, limit=None):
    """Fetch Appointed Representatives from FCA API for all major networks."""
    log.info("Phase 1.3: Fetching Appointed Representatives from FCA API...")
    output_file = RAW_DIR / "source_fca_ars.jsonl"

    if output_file.exists():
        output_file.unlink()

    # All major principal networks — ordered by estimated AR count
    NETWORKS = [
        ("190416", "St James's Place"),
        ("440703", "Quilter Financial Planning"),
        ("597301", "The Openwork Partnership"),
        ("149826", "Tenet Group"),
        ("477169", "Primis Mortgage Network"),
        ("466201", "ValidPath"),
        ("150952", "Sesame Bankhall Group"),
        ("303397", "HL Partnership"),
        ("656379", "CAERUS Capital Group"),
        ("465124", "Sense Network"),
        ("300366", "Lighthouse Group"),
        ("588184", "SimplyBiz Services"),
        ("460693", "Intrinsic Financial Services"),
        ("311584", "Fidelius Group"),
        ("475558", "Succession Group"),
        ("192590", "Sanlam Financial Services UK"),
        ("119327", "Zurich Insurance Group"),
        ("122702", "HSBC UK Bank plc"),
        ("151332", "Scottish Widows"),
        ("117355", "Aviva"),
        ("183167", "Standard Life"),
        ("181655", "Aegon"),
        ("122702", "Barclays"),
        ("141318", "Royal London"),
        ("110481", "Legal and General"),
        ("464474", "Tenet Connect Services"),
        ("186171", "Mattioli Woods"),
        ("115248", "Brewin Dolphin"),
        ("121882", "Brooks Macdonald"),
        ("213251", "Rathbones"),
    ]

    total_ars = 0
    seen_frns = set()

    for i, (frn, name) in enumerate(NETWORKS):
        log.info(f"  [{i+1}/{len(NETWORKS)}] Fetching ARs for {name} (FRN {frn})...")

        # Fetch first page
        page = 1
        while True:
            endpoint = f"Firm/{frn}/AR" if page == 1 else f"Firm/{frn}/AR?pgnp={page}"
            data = fca_api_call(endpoint)

            if not data or not isinstance(data, dict):
                break

            ar_data = data.get("Data", {})
            if not isinstance(ar_data, dict):
                break

            # FCA returns CurrentAppointedRepresentatives and PreviousAppointedRepresentatives
            current_ars = ar_data.get("CurrentAppointedRepresentatives", [])
            if not isinstance(current_ars, list):
                current_ars = []

            if not current_ars:
                break  # No current ARs or empty page

            page_count = 0
            for ar in current_ars:
                if not isinstance(ar, dict):
                    continue
                ar_frn = ar.get("FRN", "")
                ar_name = ar.get("Name", "")

                if not ar_frn or not ar_name:
                    continue
                if ar_frn in seen_frns:
                    continue

                seen_frns.add(ar_frn)
                ar_record = {
                    "frn": ar_frn,
                    "firm_name": ar_name,
                    "principal_frn": frn,
                    "principal_name": name,
                    "is_appointed_representative": True,
                    "ar_type": ar.get("Record SubType", ""),
                    "effective_date": ar.get("Effective Date", ""),
                    "tied_agent": ar.get("Tied Agent", ""),
                    "insurance_distribution": ar.get("Insurance Distribution", ""),
                    "source": "fca_api_ar",
                }
                jsonl_append(output_file, ar_record)
                total_ars += 1
                page_count += 1

            log.info(f"    Page {page}: {page_count} current ARs")

            # Check for next page
            result_info = data.get("ResultInfo", {})
            next_url = result_info.get("Next", "")
            if next_url and page_count > 0:
                page += 1
            else:
                break

        log.info(f"    Total for {name}: {total_ars} cumulative ARs")

        if limit and total_ars >= limit:
            log.info(f"  AR limit ({limit}) reached.")
            break

    log.info(f"Phase 1.3 complete: {total_ars} unique appointed representatives found across {len(NETWORKS)} networks")
    return total_ars


# ── PHASE 3: FCA API Deep Enrichment ─────────────────────────────────────────

def phase3_fca_enrich(firms, checkpoint_state, limit=None):
    """Enrich firms with full FCA API data."""
    log.info(f"Phase 3: FCA API deep enrichment ({len(firms)} firms)...")
    output_file = ENRICHED_DIR / "fca_enriched.jsonl"

    processed = set(checkpoint_state.get("processed_frns", []))
    count = 0

    for i, firm in enumerate(firms):
        frn = firm.get("frn", "")
        if not frn or frn in processed:
            continue

        if limit and count >= limit:
            log.info(f"  Test batch limit ({limit}) reached.")
            break

        log.info(f"  Enriching FRN {frn} ({count+1})... {firm.get('firm_name', '')}")

        enriched = {
            "frn": frn,
            "firm_name": firm.get("firm_name", ""),
            "source_data": firm,
        }

        # 1. Firm details — Data is a list with one dict
        firm_data = fca_api_call(f"Firm/{frn}")
        if firm_data and isinstance(firm_data, dict):
            data_list = firm_data.get("Data", [])
            fd = data_list[0] if isinstance(data_list, list) and len(data_list) > 0 else {}

            enriched["fca_firm"] = {
                "name": fd.get("Organisation Name", ""),
                "status": fd.get("Status", ""),
                "business_type": fd.get("Business Type", ""),
                "status_effective_date": fd.get("Status Effective Date", ""),
                "companies_house_number": fd.get("Companies House Number", ""),
                "client_money_permission": fd.get("Client Money Permission", ""),
                "sub_status": fd.get("Sub-Status", ""),
            }

        # 2. Permissions — Data is a dict where keys are activity names
        perms_data = fca_api_call(f"Firm/{frn}/Permissions")
        if perms_data and isinstance(perms_data, dict):
            perms_dict = perms_data.get("Data", {})
            if not isinstance(perms_dict, dict):
                perms_dict = {}

            # Activity names are the dict keys
            activity_names = list(perms_dict.keys())

            # Extract customer types and investment types from all activities
            all_customer_types = set()
            all_investment_types = set()
            for activity_name, activity_data in perms_dict.items():
                if isinstance(activity_data, list):
                    for item in activity_data:
                        if isinstance(item, dict):
                            for ct in item.get("Customer Type", []):
                                all_customer_types.add(ct)
                            for it in item.get("Investment Type", []):
                                all_investment_types.add(it)

            activities_lower = " ".join(activity_names).lower()
            types_lower = " ".join(all_customer_types).lower() + " " + " ".join(all_investment_types).lower()
            combined = activities_lower + " " + types_lower

            enriched["permissions"] = {
                "advises_on_investments": "advising on investments" in activities_lower,
                "arranges_investments": "arranging" in activities_lower,
                "manages_investments": "managing investments" in activities_lower,
                "pension_transfers": "pension transfer" in combined,
                "collective_investment_schemes": "collective investment" in combined or "unit in a collective" in combined,
                "life_policies": "life polic" in combined,
                "retail_clients": "retail" in types_lower,
                "professional_clients": "professional" in types_lower,
                "activities": activity_names[:30],
                "customer_types": sorted(all_customer_types)[:20],
                "investment_types": sorted(all_investment_types)[:30],
            }

        # 3. Individuals — Data is a list of dicts with Name, IRN, Status
        indiv_data = fca_api_call(f"Firm/{frn}/Individuals")
        if indiv_data and isinstance(indiv_data, dict):
            indivs = indiv_data.get("Data", [])
            if not isinstance(indivs, list):
                indivs = []

            enriched["individuals"] = []
            for ind in indivs[:50]:
                if isinstance(ind, dict):
                    enriched["individuals"].append({
                        "name": ind.get("Name", ""),
                        "irn": ind.get("IRN", ""),
                        "status": ind.get("Status", ""),
                    })
            result_info = indiv_data.get("ResultInfo") if indiv_data else None
            enriched["individual_count"] = int(result_info.get("total_count", len(indivs))) if result_info else len(indivs)

        # 4. Address — Data is a list of address dicts
        addr_data = fca_api_call(f"Firm/{frn}/Address")
        if addr_data and isinstance(addr_data, dict):
            addrs = addr_data.get("Data", [])
            if isinstance(addrs, list) and len(addrs) > 0:
                # Find Principal Place of Business address
                addr = addrs[0]
                for a in addrs:
                    if isinstance(a, dict) and a.get("Address Type") == "Principal Place of Business":
                        addr = a
                        break
                if isinstance(addr, dict):
                    enriched["address"] = {
                        "line1": addr.get("Address Line 1", ""),
                        "line2": addr.get("Address Line 2", ""),
                        "town": addr.get("Town", ""),
                        "county": addr.get("County", ""),
                        "postcode": addr.get("Postcode", ""),
                        "country": addr.get("Country", ""),
                    }
                    enriched["phone"] = addr.get("Phone Number", "")
                    enriched["website"] = addr.get("Website Address", "")

        # 5. Disciplinary history — Data is a list
        disc_data = fca_api_call(f"Firm/{frn}/DisciplinaryHistory")
        if disc_data and isinstance(disc_data, dict):
            disc = disc_data.get("Data", [])
            if isinstance(disc, list) and len(disc) > 0:
                enriched["has_disciplinary_history"] = True
                enriched["disciplinary_summary"] = str(disc[0])[:300] if disc else None
            else:
                enriched["has_disciplinary_history"] = False
                enriched["disciplinary_summary"] = None
        else:
            enriched["has_disciplinary_history"] = False
            enriched["disciplinary_summary"] = None

        # Write enriched record
        jsonl_append(output_file, enriched)

        # Update checkpoint
        processed.add(frn)
        count += 1

        if count % 10 == 0:
            checkpoint_state["processed_frns"] = list(processed)
            checkpoint_state["firm_count"] = count
            save_checkpoint(checkpoint_state)
            log.info(f"  Checkpoint saved: {count} firms enriched")

    # Final checkpoint
    checkpoint_state["processed_frns"] = list(processed)
    checkpoint_state["firm_count"] = count
    save_checkpoint(checkpoint_state)

    log.info(f"Phase 3 complete: {count} firms enriched via FCA API")
    return count


# ── PHASE 4: Companies House Enrichment ──────────────────────────────────────

def phase4_ch_enrich(enriched_firms, limit=None):
    """Enrich firms with Companies House data."""
    if not CH_API_KEY:
        log.warning("Phase 4: Skipping — no Companies House API key provided.")
        log.warning("  Set via: --ch-key YOUR_KEY or env var CH_API_KEY")
        return 0

    log.info(f"Phase 4: Companies House enrichment ({len(enriched_firms)} firms)...")
    output_file = ENRICHED_DIR / "ch_enriched.jsonl"

    if output_file.exists():
        output_file.unlink()

    count = 0
    for firm in enriched_firms:
        if limit and count >= limit:
            break

        firm_name = firm.get("fca_firm", {}).get("name", "") or firm.get("firm_name", "")
        postcode = firm.get("address", {}).get("postcode", "")

        if not firm_name:
            continue

        # Search Companies House by firm name
        search_query = urllib.parse.quote(firm_name)
        search_data = ch_api_call(f"/search/companies?q={search_query}&items_per_page=5")

        if not search_data or not search_data.get("items"):
            continue

        # Find best match
        best_match = None
        for item in search_data.get("items", []):
            ch_name = item.get("title", "").lower()
            search_name = firm_name.lower()

            # Simple name similarity check
            if search_name in ch_name or ch_name in search_name:
                best_match = item
                break

            # Check postcode match
            ch_addr = item.get("address", {})
            ch_postcode = ch_addr.get("postal_code", "")
            if postcode and ch_postcode and postcode[:4] == ch_postcode[:4]:
                best_match = item
                break

        if not best_match:
            continue

        ch_number = best_match.get("company_number", "")
        if not ch_number:
            continue

        # Get full company details
        company_data = ch_api_call(f"/company/{ch_number}")

        ch_record = {
            "frn": firm.get("frn", ""),
            "firm_name": firm_name,
            "companies_house_number": ch_number,
            "ch_company_name": best_match.get("title", ""),
            "ch_status": best_match.get("company_status", ""),
            "ch_type": best_match.get("company_type", ""),
            "ch_incorporated": best_match.get("date_of_creation", ""),
        }

        if company_data and isinstance(company_data, dict):
            ch_record["sic_codes"] = company_data.get("sic_codes", [])
            ch_record["registered_address"] = company_data.get("registered_office_address", {})
            accounts = company_data.get("accounts", {})
            ch_record["last_accounts_date"] = accounts.get("last_accounts", {}).get("made_up_to", "")
            ch_record["accounts_type"] = accounts.get("last_accounts", {}).get("type", "")

        # Get officers
        officers_data = ch_api_call(f"/company/{ch_number}/officers?items_per_page=20")
        if officers_data and isinstance(officers_data, dict):
            ch_record["officers"] = []
            for officer in officers_data.get("items", [])[:20]:
                if officer.get("resigned_on"):
                    continue  # Skip resigned officers
                ch_record["officers"].append({
                    "name": officer.get("name", ""),
                    "role": officer.get("officer_role", ""),
                    "appointed": officer.get("appointed_on", ""),
                })
            ch_record["active_officer_count"] = len(ch_record["officers"])

        jsonl_append(output_file, ch_record)
        count += 1

        if count % 10 == 0:
            log.info(f"  CH enriched: {count} firms matched")

    log.info(f"Phase 4 complete: {count} firms matched to Companies House")
    return count


# ── PHASE 7: Merge and Score ─────────────────────────────────────────────────

def phase7_merge_and_score():
    """Merge all enriched data and calculate scores."""
    log.info("Phase 7: Merging and scoring...")

    # Load all data
    fca_firms = jsonl_read(RAW_DIR / "source_fca_csv.jsonl")
    pimfa_firms = jsonl_read(RAW_DIR / "source_pimfa.jsonl")
    fca_enriched = jsonl_read(ENRICHED_DIR / "fca_enriched.jsonl")
    ch_enriched = jsonl_read(ENRICHED_DIR / "ch_enriched.jsonl")

    # Build lookup maps
    fca_by_frn = {f["frn"]: f for f in fca_enriched}
    ch_by_frn = {f["frn"]: f for f in ch_enriched if f.get("frn")}
    pimfa_names = {f["firm_name"].lower().strip() for f in pimfa_firms}

    output_file = ENRICHED_DIR / "ifa_master.jsonl"
    individuals_file = ENRICHED_DIR / "individuals_master.jsonl"

    if output_file.exists():
        output_file.unlink()
    if individuals_file.exists():
        individuals_file.unlink()

    # Build combined firm list: CSV + AR firms, deduplicated by FRN
    ar_firms = jsonl_read(RAW_DIR / "source_fca_ars.jsonl")
    all_firms_by_frn = {}
    for firm in fca_firms:
        all_firms_by_frn[firm["frn"]] = firm
    for ar in ar_firms:
        if ar.get("frn") and ar["frn"] not in all_firms_by_frn:
            all_firms_by_frn[ar["frn"]] = ar

    log.info(f"  Merging {len(all_firms_by_frn)} unique firms (CSV + ARs)")

    records = []
    for frn, firm in all_firms_by_frn.items():
        enriched = fca_by_frn.get(frn, {})
        ch = ch_by_frn.get(frn, {})

        # Base record
        is_ar = firm.get("is_appointed_representative", False)
        record = {
            "frn": frn,
            "firm_name": firm.get("firm_name", enriched.get("firm_name", "")),
            "status": firm.get("status", "Authorised"),
            "authorisation_date": firm.get("authorisation_date"),
            "is_appointed_representative": is_ar,
            "principal_frn": firm.get("principal_frn"),
            "principal_name": firm.get("principal_name"),
            "firm_category": "ar_practice" if is_ar else "ifa",
        }

        # FCA enrichment
        if enriched:
            fca_firm = enriched.get("fca_firm", {})
            record["phone"] = fca_firm.get("phone") or None
            record["website"] = fca_firm.get("website") or None
            record["firm_type"] = fca_firm.get("type") or None
            record["address"] = enriched.get("address", {})
            record["permissions"] = enriched.get("permissions", {})
            record["individual_count"] = enriched.get("individual_count", 0)
            record["has_disciplinary_history"] = enriched.get("has_disciplinary_history", False)
            record["disciplinary_summary"] = enriched.get("disciplinary_summary")

            # Individuals
            for ind in enriched.get("individuals", []):
                jsonl_append(individuals_file, {
                    "irn": ind.get("irn", ""),
                    "name": ind.get("name", ""),
                    "firm_frn": frn,
                    "firm_name": firm["firm_name"],
                    "role": ind.get("role", ""),
                    "status": ind.get("status", ""),
                    "start_date": ind.get("start_date", ""),
                })

            # Determine firm category from permissions
            perms = record.get("permissions", {})
            if perms.get("manages_investments"):
                record["firm_category"] = "dfm"
            elif perms.get("advises_on_investments"):
                record["firm_category"] = "ifa"

        # Companies House enrichment
        if ch:
            record["companies_house_number"] = ch.get("companies_house_number")
            record["ch_incorporated"] = ch.get("ch_incorporated")
            record["sic_codes"] = ch.get("sic_codes", [])
            record["last_accounts_date"] = ch.get("last_accounts_date")
            record["accounts_type"] = ch.get("accounts_type")
            record["active_officer_count"] = ch.get("active_officer_count", 0)

        # Pull phone and website from enriched data
        if enriched:
            record["phone"] = enriched.get("phone") or None
            record["website"] = enriched.get("website") or None
            if not record.get("phone"):
                fca_firm = enriched.get("fca_firm", {})
                record["phone"] = fca_firm.get("phone") or None
            if not record.get("website"):
                fca_firm = enriched.get("fca_firm", {})
                record["website"] = fca_firm.get("website") or None

        # PIMFA membership
        record["pimfa_member"] = record["firm_name"].lower().strip() in pimfa_names

        # Sources
        sources = []
        if frn in {f["frn"] for f in fca_firms}:
            sources.append("fca_csv")
        if firm.get("source") == "fca_api_ar":
            sources.append("fca_ar")
        if enriched:
            sources.append("fca_api")
        if ch:
            sources.append("companies_house")
        if record["pimfa_member"]:
            sources.append("pimfa")
        record["sources_found"] = sources
        record["source_count"] = len(sources)
        record["data_confidence"] = min(1.0, len(sources) * 0.25)

        # ── SCORING MODEL ──
        # Priority: retail-facing IFAs that advise on investments
        # DFMs and institutional firms score lower
        # AR practices score based on their permissions, not just AR status

        perms = record.get("permissions", {})
        individual_count = record.get("individual_count", 0)

        # 1. CORE ELIGIBILITY (0-30 points)
        # Does this firm advise retail clients on investments?
        # DFMs that only manage (don't advise retail) are deprioritised
        core_score = 0
        is_dfm = perms.get("manages_investments", False)
        is_adviser = perms.get("advises_on_investments", False)
        is_retail = perms.get("retail_clients", False)

        # DFMs are NOT distribution targets even if they also advise
        # IFAs that ONLY advise (don't manage) are the prime targets
        if is_adviser and is_retail and not is_dfm:
            core_score = 30  # Pure IFA — top priority
        elif is_adviser and is_retail and is_dfm:
            core_score = 15  # Wealth manager (advises + manages) — secondary
        elif perms.get("arranges_investments") and is_retail and not is_dfm:
            core_score = 20  # Restricted adviser — still a target
        elif is_adviser and not is_retail:
            core_score = 5   # Institutional only
        elif is_dfm and not is_adviser:
            core_score = 0   # Pure DFM — not a distribution target

        # 2. FUND DISTRIBUTION SIGNALS (0-25 points)
        fund_score = 0
        if perms.get("collective_investment_schemes"):
            fund_score += 10  # Can deal in collective investments (funds)
        if perms.get("life_policies"):
            fund_score += 5   # Life/pension wrappers often hold funds
        if perms.get("pension_transfers"):
            fund_score += 10  # Pension transfers = high value, often use multi-asset

        # 3. FIRM SIZE SIGNALS (0-20 points)
        size_score = 0
        if individual_count >= 20:
            size_score = 20  # Large firm — high distribution potential
        elif individual_count >= 10:
            size_score = 15
        elif individual_count >= 5:
            size_score = 10
        elif individual_count >= 2:
            size_score = 5
        # Solo firms score 0 — lower distribution reach

        # 4. DATA QUALITY BONUS (0-15 points)
        quality_score = 0
        if record.get("phone"):
            quality_score += 3
        if record.get("website"):
            quality_score += 3
        if record.get("address", {}).get("postcode"):
            quality_score += 2
        if record.get("pimfa_member"):
            quality_score += 5  # Industry body member = engaged firm
        if record.get("companies_house_number"):
            quality_score += 2

        # 5. DYNAMIC PLANNER PRIOR (0-6 points)
        # 40% of UK advice firms use DP, so apply 40% * 15 = 6 as prior
        dp_score = 6 if core_score >= 30 else 0  # Only for confirmed IFAs

        # 6. NEGATIVE SIGNALS
        penalty = 0
        if record.get("has_disciplinary_history"):
            penalty = -15

        # TOTAL — multiplicative: core eligibility gates everything
        # If core_score is 0 (pure DFM, not adviser), total caps at 15
        # If core_score is 30 (retail IFA), full score applies
        raw_score = core_score + fund_score + size_score + quality_score + dp_score + penalty
        if core_score >= 30:
            total_score = raw_score  # Full score for retail IFAs
        elif core_score >= 20:
            total_score = round(raw_score * 0.7)  # 70% for restricted advisers
        elif core_score >= 5:
            total_score = round(raw_score * 0.3)  # 30% for institutional-only advisers
        else:
            total_score = round(raw_score * 0.1)  # 10% for pure DFMs / non-advisers
        total_score = max(0, min(100, total_score))

        record["score_breakdown"] = {
            "core_eligibility": core_score,
            "fund_distribution": fund_score,
            "firm_size": size_score,
            "data_quality": quality_score,
            "dynamic_planner_prior": dp_score,
            "penalty": penalty,
        }
        record["overall_priority_score"] = round(total_score, 1)

        # Mandate-specific scores (adjust from base)
        record["mandate_fit_scores"] = {
            "multi_asset_cautious": total_score,
            "multi_asset_balanced": total_score,
            "multi_asset_growth": max(0, total_score - 5),
            "multi_asset_income": total_score + (5 if perms.get("pension_transfers") else 0),
            "uk_equity": max(0, total_score - 10),
            "uk_equity_income": total_score,
            "global_equity": max(0, total_score - 10),
            "corporate_bond": max(0, total_score - 5),
        }

        # Data completeness
        fields_present = sum(1 for v in [
            record.get("phone"), record.get("website"),
            record.get("address", {}).get("postcode"),
            record.get("companies_house_number"),
            record.get("individual_count"),
        ] if v)
        record["data_completeness"] = round(fields_present / 5, 2)

        record["last_updated"] = datetime.now().strftime("%Y-%m-%d")

        records.append(record)
        jsonl_append(output_file, record)

    # Assign size tiers
    records.sort(key=lambda r: r["overall_priority_score"], reverse=True)
    for i, r in enumerate(records):
        if i < 50:
            r["size_tier"] = "tier_1"
        elif i < 200:
            r["size_tier"] = "tier_2"
        elif i < 500:
            r["size_tier"] = "tier_3"
        else:
            r["size_tier"] = "tier_4"

    # Rewrite with tiers
    if output_file.exists():
        output_file.unlink()
    for r in records:
        jsonl_append(output_file, r)

    log.info(f"Phase 7 complete: {len(records)} firms in master file")
    return records


# ── Summary Report ────────────────────────────────────────────────────────────

def write_summary(records):
    """Write the run summary report."""
    summary_file = ENRICHED_DIR / "run_summary.md"

    # Stats
    total = len(records)
    with_phone = sum(1 for r in records if r.get("phone"))
    with_website = sum(1 for r in records if r.get("website"))
    with_ch = sum(1 for r in records if r.get("companies_house_number"))
    with_perms = sum(1 for r in records if r.get("permissions"))
    pimfa = sum(1 for r in records if r.get("pimfa_member"))
    disciplinary = sum(1 for r in records if r.get("has_disciplinary_history"))

    # Category breakdown
    categories = {}
    for r in records:
        cat = r.get("firm_category", "unknown")
        categories[cat] = categories.get(cat, 0) + 1

    # Tier breakdown
    tiers = {}
    for r in records:
        tier = r.get("size_tier", "unassigned")
        tiers[tier] = tiers.get(tier, 0) + 1

    # Region breakdown
    regions = {}
    for r in records:
        pc = r.get("address", {}).get("postcode", "")
        if pc:
            region = pc[:2].rstrip("0123456789")
            regions[region] = regions.get(region, 0) + 1

    # Top 20
    top20 = sorted(records, key=lambda r: r.get("overall_priority_score", 0), reverse=True)[:20]

    with open(summary_file, "w") as f:
        f.write("# IFA Intelligence Pipeline — Run Summary\n\n")
        f.write(f"**Run date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"**Total firms:** {total}\n\n")

        f.write("## Universe\n\n")
        f.write(f"| Source | Count |\n|--------|-------|\n")
        f.write(f"| FCA CSV (after filtering) | {total} |\n")
        f.write(f"| PIMFA members matched | {pimfa} |\n\n")

        f.write("## Enrichment Completion\n\n")
        f.write(f"| Field | Count | % |\n|-------|-------|---|\n")
        f.write(f"| FCA API enriched | {with_perms} | {with_perms*100//max(total,1)}% |\n")
        f.write(f"| Companies House matched | {with_ch} | {with_ch*100//max(total,1)}% |\n")
        f.write(f"| Has phone | {with_phone} | {with_phone*100//max(total,1)}% |\n")
        f.write(f"| Has website | {with_website} | {with_website*100//max(total,1)}% |\n\n")

        f.write("## Category Breakdown\n\n")
        f.write(f"| Category | Count |\n|----------|-------|\n")
        for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
            f.write(f"| {cat} | {cnt} |\n")

        f.write(f"\n## Tier Breakdown\n\n")
        f.write(f"| Tier | Count |\n|------|-------|\n")
        for tier in ["tier_1", "tier_2", "tier_3", "tier_4"]:
            f.write(f"| {tier} | {tiers.get(tier, 0)} |\n")

        f.write(f"\n## Top 20 Firms by Priority Score\n\n")
        f.write(f"| Rank | Firm | FRN | Category | Score | Sources |\n")
        f.write(f"|------|------|-----|----------|-------|--------|\n")
        for i, r in enumerate(top20):
            f.write(f"| {i+1} | {r['firm_name'][:40]} | {r['frn']} | {r.get('firm_category','')} | {r.get('overall_priority_score',0)} | {r.get('source_count',0)} |\n")

        f.write(f"\n## Data Quality\n\n")
        f.write(f"- Firms with disciplinary history: {disciplinary}\n")
        f.write(f"- Firms with no phone or website: {total - with_phone - with_website + sum(1 for r in records if r.get('phone') and r.get('website'))}\n")
        f.write(f"- Firms with zero individuals: {sum(1 for r in records if r.get('individual_count', 0) == 0)}\n")

        f.write(f"\n## Recommended Next Steps\n\n")
        f.write(f"1. Run full pipeline (all {total} firms) with: `python3 pipeline.py --full`\n")
        f.write(f"2. Add Companies House API key for CH enrichment\n")
        f.write(f"3. Add VouchedFor/Unbiased web scraping (Phases not yet implemented)\n")
        f.write(f"4. Add website intelligence scraping (Phase 5)\n")
        f.write(f"5. Add news signal detection (Phase 6)\n")

    log.info(f"Summary written to {summary_file}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="IFA Intelligence Pipeline")
    parser.add_argument("--test", action="store_true", help="Run 100-firm test batch")
    parser.add_argument("--full", action="store_true", help="Run full pipeline")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint")
    parser.add_argument("--phase", type=int, help="Run specific phase only")
    parser.add_argument("--ch-key", type=str, help="Companies House API key")
    args = parser.parse_args()

    if args.ch_key:
        global CH_API_KEY
        CH_API_KEY = args.ch_key

    limit = TEST_BATCH_SIZE if args.test else None
    checkpoint = load_checkpoint()

    log.info("=" * 60)
    log.info("IFA Intelligence Pipeline")
    log.info(f"Mode: {'TEST (100 firms)' if args.test else 'FULL'}")
    log.info(f"Output: {BASE_DIR}")
    log.info("=" * 60)

    start_time = datetime.now()

    # Phase 1: Universe Discovery
    if not args.phase or args.phase == 1:
        fca_firms = phase1_fca_csv()
        pimfa_firms = phase1_pimfa()

        # AR discovery (limited for test batch)
        if args.test:
            phase1_fca_api_ars(fca_firms, limit=50)
        else:
            phase1_fca_api_ars(fca_firms)

        checkpoint["phase_status"]["phase1"] = "complete"
        save_checkpoint(checkpoint)

    # Load all firms for subsequent phases — combine CSV + ARs
    fca_firms = jsonl_read(RAW_DIR / "source_fca_csv.jsonl")
    ar_firms = jsonl_read(RAW_DIR / "source_fca_ars.jsonl")

    # Deduplicate: AR firms may overlap with CSV firms by FRN
    seen_frns = {f["frn"] for f in fca_firms if f.get("frn")}
    new_ar_firms = 0
    for ar in ar_firms:
        if ar.get("frn") and ar["frn"] not in seen_frns:
            seen_frns.add(ar["frn"])
            fca_firms.append(ar)
            new_ar_firms += 1

    log.info(f"Combined universe: {len(fca_firms)} firms ({len(fca_firms) - new_ar_firms} CSV + {new_ar_firms} new from ARs)")

    # Apply test batch limit
    if limit:
        fca_firms = fca_firms[:limit]
        log.info(f"Test batch: processing first {limit} firms")

    # Phase 3: FCA API Deep Enrichment
    if not args.phase or args.phase == 3:
        phase3_fca_enrich(fca_firms, checkpoint, limit=limit)
        checkpoint["phase_status"]["phase3"] = "complete"
        save_checkpoint(checkpoint)

    # Phase 4: Companies House
    if not args.phase or args.phase == 4:
        enriched = jsonl_read(ENRICHED_DIR / "fca_enriched.jsonl")
        phase4_ch_enrich(enriched, limit=limit)
        checkpoint["phase_status"]["phase4"] = "complete"
        save_checkpoint(checkpoint)

    # Phase 7: Merge and Score
    if not args.phase or args.phase == 7:
        records = phase7_merge_and_score()
        write_summary(records)
        checkpoint["phase_status"]["phase7"] = "complete"
        save_checkpoint(checkpoint)

    elapsed = (datetime.now() - start_time).total_seconds()
    log.info(f"Pipeline complete in {elapsed:.0f}s ({elapsed/60:.1f} minutes)")
    log.info(f"Output files in: {ENRICHED_DIR}")


if __name__ == "__main__":
    main()
